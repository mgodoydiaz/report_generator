"""Router /api/tables — CRUD de tablas configurables (B7).

Cada tabla vive como un Spec con type='Tablas'. El campo `tables_list`
del Spec contiene un array con UN único TableConfig (1 spec = 1 tabla).
Esto mantiene la simetría con specs de Reportes/Gráficos.

Endpoints:
    GET    /api/tables/                 lista resumen (sidebar)
    POST   /api/tables/                 crear
    GET    /api/tables/{id}             leer detalle (config completa)
    PUT    /api/tables/{id}             actualizar
    DELETE /api/tables/{id}             borrar
    POST   /api/tables/{id}/duplicate   clonar
    GET    /api/tables/{id}/data        preview con datos reales formateados
"""
from __future__ import annotations

import json
import threading
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy import event
from sqlalchemy.orm import Session

from backend.auth import get_current_user, require_editor
from backend.database import get_db
from backend.logging_config import get_logger
from backend.models import Indicator, Metric, MetricData, MetricDimension, Dimension, Spec, User
from backend.rgenerator.core.pivot_engine import pivot
from backend.schemas_table import TableConfig, TableCreate, TableSummary, TableUpdate

logger = get_logger(__name__)

router = APIRouter(prefix="/api/tables", tags=["tables"])


SPEC_TYPE = "Tablas"


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────


def _now_str() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def _parse_meta(spec: Spec) -> Dict[str, Any]:
    try:
        return json.loads(spec.metadata_ or "{}")
    except Exception:
        return {}


def _parse_tables_list(spec: Spec) -> List[Dict[str, Any]]:
    try:
        return json.loads(spec.tables_list or "[]")
    except Exception:
        return []


def _spec_to_summary(spec: Spec) -> TableSummary:
    meta = _parse_meta(spec)
    tables = _parse_tables_list(spec)
    cfg = tables[0] if tables else {}
    cols = cfg.get("columns") or []
    ds = cfg.get("data_source") or {}
    return TableSummary(
        id_spec=spec.id_spec,
        name=spec.name,
        description=meta.get("description", ""),
        is_draft=meta.get("is_draft", True),
        metric_id=ds.get("metric_id"),
        n_columns=len(cols),
        updated_at=meta.get("updated_at", ""),
    )


def _spec_to_full(spec: Spec) -> Dict[str, Any]:
    meta = _parse_meta(spec)
    tables = _parse_tables_list(spec)
    return {
        "id_spec": spec.id_spec,
        "name": spec.name,
        "description": meta.get("description", ""),
        "is_draft": meta.get("is_draft", True),
        "updated_at": meta.get("updated_at", ""),
        "config": tables[0] if tables else None,
    }


def _get_spec_or_404(db: Session, table_id: int, org_id: int) -> Spec:
    spec = db.query(Spec).filter(
        Spec.id_spec == table_id,
        Spec.org_id == org_id,
        Spec.type == SPEC_TYPE,
    ).first()
    if not spec:
        raise HTTPException(status_code=404, detail=f"Tabla {table_id} no encontrada")
    return spec


# ─────────────────────────────────────────────────────────────────────────
# Carga de datos (compartido con LoadMetricToDF, aplicado al preview)
# ─────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────
# Cache TTL para _load_metric_to_df
# ─────────────────────────────────────────────────────────────────────────
#
# Un dashboard llena ~10 charts/tables, muchos de ellos sobre la misma
# métrica. Sin cache cada uno repite las queries y el parsing de JSON.
# Con TTL corto el segundo+ load del mismo (metric, filtros) es instantáneo.
# El TTL de 60s acota la ventana de obsolescencia tras un ETL.
_METRIC_DF_CACHE: Dict[Tuple, Tuple[float, pd.DataFrame]] = {}
_METRIC_DF_CACHE_LOCK = threading.Lock()
_METRIC_DF_CACHE_TTL = 60.0


def _metric_df_cache_key(org_id: int, metric_id: int,
                         filters: Optional[Dict[str, Any]]) -> Tuple:
    if not filters:
        return (org_id, metric_id, None)
    items = []
    for k in sorted(filters.keys()):
        v = filters[k]
        if isinstance(v, (list, tuple, set)):
            v = tuple(sorted(str(x) for x in v))
        items.append((str(k), v))
    return (org_id, metric_id, tuple(items))


def invalidate_metric_df_cache(metric_id: Optional[int] = None) -> None:
    """Invalida entradas del cache. Si metric_id es None, limpia todo.

    Llamar desde endpoints que escriben MetricData (carga ETL, edición de
    valores, etc.) para que el siguiente read vea los nuevos datos sin
    esperar al TTL.
    """
    with _METRIC_DF_CACHE_LOCK:
        if metric_id is None:
            _METRIC_DF_CACHE.clear()
            return
        keys = [k for k in _METRIC_DF_CACHE if k[1] == metric_id]
        for k in keys:
            _METRIC_DF_CACHE.pop(k, None)


def _load_metric_to_df(db: Session, org_id: int, metric_id: int,
                       filters: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    """Wrapper cacheado de _load_metric_to_df_uncached. Devuelve una copia
    del DataFrame para que el caller pueda mutarla sin afectar el cache.
    """
    key = _metric_df_cache_key(org_id, metric_id, filters)
    now = time.time()
    with _METRIC_DF_CACHE_LOCK:
        cached = _METRIC_DF_CACHE.get(key)
        if cached and (now - cached[0]) < _METRIC_DF_CACHE_TTL:
            return cached[1].copy()

    df = _load_metric_to_df_uncached(db, org_id, metric_id, filters)

    with _METRIC_DF_CACHE_LOCK:
        _METRIC_DF_CACHE[key] = (now, df)
        # Cleanup de entradas expiradas (>2x TTL) para evitar memory leak
        cutoff = now - 2 * _METRIC_DF_CACHE_TTL
        expired = [k for k, (t, _) in _METRIC_DF_CACHE.items() if t < cutoff]
        for k in expired:
            _METRIC_DF_CACHE.pop(k, None)

    return df.copy()


def _load_metric_to_df_uncached(db: Session, org_id: int, metric_id: int,
                                filters: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    """Carga metric_data + dimensiones a DataFrame plano, aplicando filtros
    por igualdad simple sobre nombres de dimensiones.

    Es una versión inline de la lógica que vive en
    `rgenerator/core/metric_steps.py:LoadMetricToDF` — replicada acá
    para que el endpoint /data no dependa de instanciar steps.
    """
    metric = db.query(Metric).filter(
        Metric.id_metric == metric_id,
        Metric.org_id == org_id,
    ).first()
    if not metric:
        raise HTTPException(status_code=404, detail=f"Métrica {metric_id} no encontrada")

    dim_links = db.query(MetricDimension).filter(MetricDimension.id_metric == metric_id).all()
    dim_ids = [lnk.id_dimension for lnk in dim_links]
    dims = db.query(Dimension).filter(Dimension.id_dimension.in_(dim_ids)).all() if dim_ids else []
    dims_map = {d.id_dimension: d.name for d in dims}

    rows = db.query(MetricData).filter(MetricData.id_metric == metric_id).all()

    # Parse de meta_json UNA sola vez fuera del loop: es invariante por métrica
    # y antes se parseaba O(N) veces (una por row).
    try:
        meta = json.loads(metric.meta_json or "{}") if isinstance(metric.meta_json, str) else (metric.meta_json or {})
    except Exception:
        meta = {}
    meta_fields = meta.get("fields", []) if isinstance(meta, dict) else []

    flat: List[Dict[str, Any]] = []
    for r in rows:
        item: Dict[str, Any] = {}
        # Dimensiones
        try:
            dims_json = json.loads(r.dimensions_json) if isinstance(r.dimensions_json, str) else (r.dimensions_json or {})
        except Exception:
            dims_json = {}
        for dim_id, name in dims_map.items():
            item[name] = dims_json.get(str(dim_id))
        # Valor (object → expandido a fields, simple → 1 columna)
        val = r.value
        if metric.data_type == "object":
            try:
                val_obj = json.loads(val) if isinstance(val, str) else val
            except Exception:
                val_obj = {}
            for f in meta_fields:
                fname = f["name"]
                item[fname] = val_obj.get(fname)
        else:
            try:
                if metric.data_type == "int":
                    item[metric.name] = int(val)
                elif metric.data_type == "float":
                    item[metric.name] = float(val)
                else:
                    item[metric.name] = val
            except Exception:
                item[metric.name] = val
        flat.append(item)

    df = pd.DataFrame(flat)

    # Filtros: igualdad simple (str) o IN (list de valores).
    # Soporta multi-valor desde B9: cuando val es list/tuple, hace
    # df[col].isin([...]) para retener cualquier coincidencia.
    if filters:
        for col, val in filters.items():
            if col not in df.columns:
                continue
            if isinstance(val, (list, tuple, set)):
                allowed = {str(v) for v in val}
                if not allowed:
                    continue
                df = df[df[col].astype(str).isin(allowed)]
            else:
                df = df[df[col].astype(str) == str(val)]

    return df


def _apply_format(value: Any, fmt: str, decimals: int = 1) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if fmt == "int":
        try:
            return f"{int(value)}"
        except (ValueError, TypeError):
            return str(value)
    if fmt == "float":
        try:
            return f"{float(value):.{decimals}f}"
        except (ValueError, TypeError):
            return str(value)
    if fmt == "percent":
        try:
            return f"{float(value) * 100:.{decimals}f}%"
        except (ValueError, TypeError):
            return str(value)
    if fmt == "date":
        return str(value)
    return str(value)


def _resolve_color_for_value(value: Any, color_scale: Dict[str, Any],
                              row: Dict[str, Any], indicator_levels_cache: Dict[int, list]) -> Optional[str]:
    """Devuelve color hex para una celda según el color_scale."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    kind = color_scale.get("kind")
    if kind == "linked_indicator":
        ind_id = color_scale.get("indicator_id")
        level_field = color_scale.get("level_field")
        levels = indicator_levels_cache.get(ind_id, [])
        if not level_field or not levels:
            return None
        level_name = row.get(level_field)
        if not level_name:
            return None
        for lvl in levels:
            if str(lvl.get("name", "")).lower() == str(level_name).lower():
                return lvl.get("color")
    elif kind == "diverging":
        try:
            v = float(value)
        except (ValueError, TypeError):
            return None
        mp = float(color_scale.get("midpoint", 0))
        if v < mp:
            return color_scale.get("min_color")
        if v > mp:
            return color_scale.get("max_color")
        return color_scale.get("neutral_color")
    elif kind == "sequential":
        return color_scale.get("base_color")
    return None


def _load_indicator_levels(db: Session, org_id: int, indicator_ids: List[int]) -> Dict[int, list]:
    """Cache: {id_indicator: [{name, color, order}]}."""
    if not indicator_ids:
        return {}
    inds = db.query(Indicator).filter(
        Indicator.id_indicator.in_(indicator_ids),
        Indicator.org_id == org_id,
    ).all()
    out: Dict[int, list] = {}
    for ind in inds:
        try:
            levels = json.loads(ind.achievement_levels or "[]")
        except Exception:
            levels = []
        out[ind.id_indicator] = levels
    return out


# ─────────────────────────────────────────────────────────────────────────
# CRUD endpoints
# ─────────────────────────────────────────────────────────────────────────


@router.get("/", response_model=List[TableSummary])
def list_tables(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    specs = db.query(Spec).filter(Spec.org_id == user.org_id, Spec.type == SPEC_TYPE).order_by(Spec.id_spec.desc()).all()
    return [_spec_to_summary(s) for s in specs]


@router.post("/")
def create_table(
    payload: TableCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_editor),
):
    meta = {
        "description": payload.description,
        "is_draft": payload.is_draft,
        "updated_at": _now_str(),
    }
    spec = Spec(
        name=payload.name,
        type=SPEC_TYPE,
        metadata_=json.dumps(meta, ensure_ascii=False),
        charts_list="[]",
        tables_list=json.dumps([payload.config.model_dump()], ensure_ascii=False),
        org_id=user.org_id,
    )
    db.add(spec)
    db.commit()
    db.refresh(spec)
    return {"status": "success", "id_spec": spec.id_spec}


@router.get("/{table_id}")
def get_table(
    table_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    spec = _get_spec_or_404(db, table_id, user.org_id)
    return _spec_to_full(spec)


@router.put("/{table_id}")
def update_table(
    table_id: int,
    payload: TableUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_editor),
):
    spec = _get_spec_or_404(db, table_id, user.org_id)
    meta = _parse_meta(spec)
    if payload.name is not None:
        spec.name = payload.name
    if payload.description is not None:
        meta["description"] = payload.description
    if payload.is_draft is not None:
        meta["is_draft"] = payload.is_draft
    if payload.config is not None:
        spec.tables_list = json.dumps([payload.config.model_dump()], ensure_ascii=False)
    meta["updated_at"] = _now_str()
    spec.metadata_ = json.dumps(meta, ensure_ascii=False)
    db.commit()
    db.refresh(spec)
    return {"status": "success", "id_spec": spec.id_spec}


@router.delete("/{table_id}")
def delete_table(
    table_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_editor),
):
    spec = _get_spec_or_404(db, table_id, user.org_id)
    db.delete(spec)
    db.commit()
    return {"status": "success", "message": f"Tabla {table_id} eliminada"}


@router.post("/{table_id}/duplicate")
def duplicate_table(
    table_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_editor),
):
    spec = _get_spec_or_404(db, table_id, user.org_id)
    new = Spec(
        name=spec.name + " (Copia)",
        type=spec.type,
        metadata_=spec.metadata_,
        charts_list=spec.charts_list,
        tables_list=spec.tables_list,
        org_id=spec.org_id,
    )
    db.add(new)
    db.commit()
    db.refresh(new)
    return {"status": "success", "id_spec": new.id_spec}


# ─────────────────────────────────────────────────────────────────────────
# Preview de datos
# ─────────────────────────────────────────────────────────────────────────


def _prepare_table_df(
    db: Session, org_id: int, cfg: TableConfig,
    extra_filters: Optional[Dict[str, Any]] = None,
) -> pd.DataFrame:
    """Carga + prepara el DataFrame de una TableConfig SIN grouping/sort/format.

    Aplica: merge de filtros, herencia de derived_columns del indicador,
    carga de la métrica con pre-filtros no-temporales, cálculo de
    derived_fields y filtros temporales post-cálculo. Es la base común del
    render tabular clásico (`_render_table_data`) y del modo pivote
    (dashboard `/data` + export Excel), garantizando que ambos vean el
    mismo df de partida (una sola fuente de verdad).
    """
    base_filters = dict(cfg.data_source.filters)
    if extra_filters:
        base_filters.update(extra_filters)

    # Carga inicial SIN aplicar filtros temporales — los derived_fields
    # (slope/delta) necesitan ver TODOS los puntos del estudiante para
    # calcular correctamente. Se identifica qué dims son temporales
    # leyendo `temporal_dim_ids` del derived_fields_override (o de los
    # derived_columns del indicador linked, ver más abajo).
    derived_cfg_list = list(cfg.data_source.derived_fields_override or [])

    # Si el spec NO trae derived_fields_override, intentamos heredar del
    # indicador linked a la metric (single source of truth: el spec del
    # indicador). Solo se aplica si el indicador tiene derived_columns y
    # apunta a la misma metric.
    if not derived_cfg_list:
        from backend.models import IndicatorMetric
        ind_links = db.query(IndicatorMetric).filter(
            IndicatorMetric.id_metric == cfg.data_source.metric_id
        ).all()
        for lnk in ind_links:
            ind = db.query(Indicator).filter(
                Indicator.id_indicator == lnk.id_indicator,
                Indicator.org_id == org_id,
            ).first()
            if not ind or not ind.derived_columns:
                continue
            try:
                ind_dc = json.loads(ind.derived_columns)
            except Exception:
                continue
            for entry in ind_dc:
                if entry.get("metric_id") == cfg.data_source.metric_id:
                    derived_cfg_list.append(entry)

    # Identifica las dims temporales para excluirlas del filtro pre-cálculo
    temporal_dim_ids: set[str] = set()
    for entry in derived_cfg_list:
        for did in (entry.get("temporal_dim_ids") or []):
            temporal_dim_ids.add(str(did))
    # Resolver nombres de dim temporales (ej "Mes", "N Prueba")
    temporal_dim_names: set[str] = set()
    if temporal_dim_ids:
        dims = db.query(Dimension).filter(
            Dimension.id_dimension.in_([int(x) for x in temporal_dim_ids])
        ).all()
        temporal_dim_names = {d.name for d in dims}

    # Pre-filtros: solo los NO-temporales se aplican antes del cálculo
    pre_filters = {k: v for k, v in base_filters.items() if k not in temporal_dim_names}
    df = _load_metric_to_df(db, org_id, cfg.data_source.metric_id, pre_filters)

    # Aplicar derived_columns sobre el df ANTES de filtrar por temporal
    if derived_cfg_list and not df.empty:
        try:
            from backend.rgenerator.core.derived_fields_engine import apply_derived_fields
            for entry in derived_cfg_list:
                configs = entry.get("configs") or []
                if configs:
                    df = apply_derived_fields(df, configs)
        except Exception:
            logger.error("Error aplicando derived_columns en preview de tabla", exc_info=True)

    # Aplicar filtros temporales POST cálculo
    post_temporal = {k: v for k, v in base_filters.items() if k in temporal_dim_names}
    if post_temporal:
        for col, val in post_temporal.items():
            if col not in df.columns:
                continue
            if isinstance(val, (list, tuple, set)):
                allowed = {str(v) for v in val}
                if not allowed:
                    continue
                df = df[df[col].astype(str).isin(allowed)]
            else:
                df = df[df[col].astype(str) == str(val)]

    return df


def _render_table_data(
    db: Session, org_id: int, cfg: TableConfig,
    limit: int, offset: int, include_styles: bool,
    extra_filters: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Aplica el pipeline completo (filtros → grouping → sort → format → color)
    sobre una TableConfig y devuelve la respuesta estandar `{columns, rows, total_rows, limit, offset}`.

    Compartido por GET /{id}/data (config persistida) y POST /preview
    (config draft sin persistir, pensado para el editor live).
    """
    df = _prepare_table_df(db, org_id, cfg, extra_filters)

    # Grouping con multi-agg sobre la misma columna fuente.
    # Cada TableColumn con `agg` se vuelve un NamedAgg(column=source_key,
    # aggfunc=agg) cuyo alias en el df resultante es la `key` de la
    # columna. Eso permite que dos columnas con `source_key="Logro"` y
    # distinto `agg` produzcan dos columnas separadas en el output (ej
    # "Logro_mean", "Logro_max"). Si no se pasó source_key, source = key
    # → comportamiento 1-a-1 anterior.
    if cfg.behavior.grouping:
        gb_list = cfg.behavior.grouping.by_list()
        gb_present = [g for g in gb_list if g in df.columns]
        if gb_present:
            named_aggs: Dict[str, Any] = {}
            for c in cfg.columns:
                if c.key in gb_present:
                    continue
                src = c.resolved_source_key()
                if src not in df.columns:
                    continue
                if c.agg:
                    named_aggs[c.key] = pd.NamedAgg(column=src, aggfunc=c.agg)
            if named_aggs:
                # Cuando hay >1 col en groupby pandas devuelve un MultiIndex
                # tupla; `as_index=False` lo aplana a columnas regulares.
                df = df.groupby(gb_present, as_index=False).agg(**named_aggs)

    # Sort
    for s in cfg.behavior.sorting:
        if s.column in df.columns:
            df = df.sort_values(by=s.column, ascending=(s.dir == "asc"))

    total_rows = len(df)
    df_page = df.iloc[offset: offset + limit].copy()

    # Color scales — pre-cargar achievement_levels de indicadores referenciados
    indicator_ids = [
        c.color_scale.indicator_id for c in cfg.columns
        if c.color_scale and c.color_scale.kind == "linked_indicator"
    ]
    indicator_levels_cache = _load_indicator_levels(db, org_id, list(set(indicator_ids))) if indicator_ids else {}

    columns_meta = [
        {"key": c.key, "header": c.header, "format": c.format,
         "pinned": c.pinned, "width": c.width}
        for c in cfg.columns if not c.hidden
    ]

    # Después del groupby con NamedAgg, las columnas resultantes se llaman
    # como las `key` aliased. Si no hubo groupby, df conserva las columnas
    # originales del metric_data, identificadas por `source_key`.
    rows_out = []
    for _, row in df_page.iterrows():
        row_obj: Dict[str, Any] = {}
        for c in cfg.columns:
            if c.hidden:
                continue
            # Resolver el campo real en el df actual (post-groupby o pre).
            if c.key in df_page.columns:
                lookup = c.key
            else:
                src = c.resolved_source_key()
                lookup = src if src in df_page.columns else None
            raw = row.get(lookup) if lookup else None
            # value_aliases se aplica al raw antes de _apply_format. Esto
            # cambia solo `formatted` — `raw` queda intacto para filtros,
            # exports CSV/XLSX y comparaciones.
            display = raw
            if c.value_aliases and raw is not None:
                key_str = str(raw)
                if key_str in c.value_aliases:
                    display = c.value_aliases[key_str]
            cell: Dict[str, Any] = {
                "raw": None if (isinstance(raw, float) and pd.isna(raw)) else raw,
                "formatted": _apply_format(display, c.format, c.decimals),
            }
            if include_styles and c.color_scale:
                color = _resolve_color_for_value(raw, c.color_scale.model_dump(), row.to_dict(), indicator_levels_cache)
                if color:
                    cell["color"] = color
            row_obj[c.key] = cell
        rows_out.append(row_obj)

    return {
        "columns": columns_meta,
        "rows": rows_out,
        "total_rows": total_rows,
        "limit": limit,
        "offset": offset,
    }


# ─────────────────────────────────────────────────────────────────────────
# Modo pivote (W2) — dashboard PivotResult JSON + export Excel
# ─────────────────────────────────────────────────────────────────────────


def _render_pivot_data(
    db: Session, org_id: int, cfg: TableConfig,
    extra_filters: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Corre el motor de pivotes sobre el df de la tabla y devuelve el
    `PivotResult` serializado a JSON para el dashboard (PARTE B2 frontend).

    Forma de la respuesta::

        {"mode": "pivot",
         "pivot": <PivotResult.model_dump(mode="json")>,
         "n_rows": <filas del df de origen>}
    """
    df = _prepare_table_df(db, org_id, cfg, extra_filters)
    result = pivot(df, cfg.pivot)
    return {
        "mode": "pivot",
        "pivot": result.model_dump(mode="json"),
        "n_rows": int(len(df)),
    }


def _excel_number_format(fmt: Optional[str], agg: str) -> str:
    """Traduce el `format` de una métrica (Python format-spec) al
    `number_format` de openpyxl. Los valores crudos van sin formatear en la
    celda; el number_format controla su presentación en Excel."""
    if fmt:
        f = fmt.strip()
        if f.endswith("%"):
            # ".1%" → 1 decimal; ".0%"/"%" → entero
            decimals = 0
            dot = f.rfind(".")
            if dot != -1 and dot + 1 < len(f) - 1:
                try:
                    decimals = int(f[dot + 1:-1])
                except ValueError:
                    decimals = 0
            return "0." + ("0" * decimals) + "%" if decimals else "0%"
        if f.endswith("f"):
            decimals = 0
            dot = f.rfind(".")
            if dot != -1:
                try:
                    decimals = int(f[dot + 1:-1])
                except ValueError:
                    decimals = 0
            return "0." + ("0" * decimals) if decimals else "0"
        if f.endswith(("d",)):
            return "0"
    # Defaults por agregación cuando no hay format explícito.
    from backend.schemas_pivot import PCT_AGGS
    if agg in PCT_AGGS:
        return "0.0%"
    if agg in ("count", "nunique"):
        return "0"
    return "General"


def _pivot_to_xlsx_bytes(cfg: TableConfig, result) -> bytes:
    """Serializa un `PivotResult` a un .xlsx (openpyxl) con valores crudos y
    `number_format` derivado del `format` de cada métrica del spec."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Pivote"

    total_label = result.meta.total_label
    # Formato por (field, agg) desde meta.aggs (una métrica del spec).
    fmt_by_metric: Dict[tuple, str] = {}
    for a in result.meta.aggs:
        fmt_by_metric[(a["field"], a["agg"])] = _excel_number_format(a.get("format"), a["agg"])

    # Encabezados: campos de fila + una columna por PivotColumn.
    headers: List[str] = list(result.row_fields)
    col_number_formats: List[Optional[str]] = [None] * len(result.row_fields)
    for col in result.columns:
        level = " · ".join(col.keys) if col.keys else ""
        headers.append(f"{level} · {col.label}" if level else col.label)
        col_number_formats.append(fmt_by_metric.get((col.field, col.agg), "General"))
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    n_row_fields = len(result.row_fields)
    for prow in result.rows:
        values: List[Any] = []
        for i in range(n_row_fields):
            if prow.is_total:
                values.append(total_label if i == 0 else "")
            else:
                values.append(prow.keys[i] if i < len(prow.keys) else "")
        for pcell in prow.cells:
            values.append(pcell.value)  # número crudo o None
        ws.append(values)

    # Aplicar number_format a las columnas de datos (post-header).
    for col_idx, nf in enumerate(col_number_formats, start=1):
        if not nf or nf == "General":
            continue
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = nf

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{table_id}/data")
def get_table_data(
    table_id: int,
    limit: int = Query(50, ge=1, le=2000),
    offset: int = Query(0, ge=0),
    include_styles: bool = Query(True),
    extra_filters: Optional[str] = Query(None, description="JSON dict con filtros adicionales (encoded)"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Devuelve datos formateados de una tabla persistida."""
    spec = _get_spec_or_404(db, table_id, user.org_id)
    tables = _parse_tables_list(spec)
    if not tables:
        raise HTTPException(status_code=400, detail="Tabla sin configuración")
    try:
        cfg = TableConfig(**tables[0])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Config inválida: {e}")

    extra: Optional[Dict[str, Any]] = None
    if extra_filters:
        try:
            extra = json.loads(extra_filters)
        except Exception:
            extra = None
    # Modo pivote (W2): devuelve el PivotResult calculado por el motor.
    if cfg.pivot is not None:
        try:
            return _render_pivot_data(db, user.org_id, cfg, extra)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    return _render_table_data(db, user.org_id, cfg, limit, offset, include_styles, extra)


class TablePreviewRequest(BaseModel):
    config: TableConfig
    limit: int = 50
    offset: int = 0
    include_styles: bool = True
    extra_filters: Optional[Dict[str, Any]] = None


@router.post("/preview")
def preview_table_config(
    payload: "TablePreviewRequest",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Igual que /{id}/data pero recibe la config en body (no requiere
    que la tabla esté persistida). Pensado para el editor live."""
    if payload.config.pivot is not None:
        try:
            return _render_pivot_data(db, user.org_id, payload.config, payload.extra_filters)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    return _render_table_data(
        db, user.org_id, payload.config,
        payload.limit, payload.offset, payload.include_styles,
        payload.extra_filters,
    )


@router.get("/{table_id}/export-pivot")
def export_pivot_xlsx(
    table_id: int,
    extra_filters: Optional[str] = Query(None, description="JSON dict con filtros adicionales (encoded)"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Exporta una tabla en modo pivote a un archivo .xlsx descargable.

    Corre el motor de pivotes sobre el df de la tabla y escribe los valores
    crudos con `number_format` derivado del `format` del spec. Auth JWT +
    multi-tenant por org_id (404 si la tabla es de otra org).
    """
    from fastapi.responses import Response

    spec = _get_spec_or_404(db, table_id, user.org_id)
    tables = _parse_tables_list(spec)
    if not tables:
        raise HTTPException(status_code=400, detail="Tabla sin configuración")
    try:
        cfg = TableConfig(**tables[0])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Config inválida: {e}")
    if cfg.pivot is None:
        raise HTTPException(status_code=400, detail="La tabla no está en modo pivote")

    extra: Optional[Dict[str, Any]] = None
    if extra_filters:
        try:
            extra = json.loads(extra_filters)
        except Exception:
            extra = None

    df = _prepare_table_df(db, user.org_id, cfg, extra)
    try:
        result = pivot(df, cfg.pivot)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    xlsx_bytes = _pivot_to_xlsx_bytes(cfg, result)

    filename = f"pivote_{table_id}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────
# Invalidación automática del cache cuando cambia MetricData
# ─────────────────────────────────────────────────────────────────────────
#
# Cubre los endpoints HTTP de metrics, los steps de pipelines (SaveToMetric)
# y cualquier otra escritura que use la sesión SQLAlchemy. Garantiza que el
# dashboard refleje los cambios sin esperar el TTL.
@event.listens_for(MetricData, "after_insert")
@event.listens_for(MetricData, "after_update")
@event.listens_for(MetricData, "after_delete")
def _invalidate_on_metric_data_change(mapper, connection, target):
    invalidate_metric_df_cache(target.id_metric)
