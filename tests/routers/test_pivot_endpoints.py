"""Tests de integración de los consumidores del motor de pivotes (W2-B1).

Cubre los tres consumidores backend cableados al motor puro
(`backend.rgenerator.core.pivot_engine`):

- **Dashboard**: `GET /api/tables/{id}/data` de una tabla en modo pivote
  devuelve el `PivotResult` serializado a JSON.
- **Excel**: `GET /api/tables/{id}/export-pivot` produce un .xlsx válido con
  los valores crudos y number_format derivado del format del spec.
- **PDF v2**: `runtime` con una sección `pivot` genera bytes sin `{{`
  residual y la sección presente.

Además:
- **Paridad**: la MISMA PivotSpec sobre el mismo df da los mismos números en
  los tres consumidores (garantía de una sola fuente de verdad).
- **Multi-tenant**: los endpoints nuevos respetan org_id (404 cross-org).

Ver docs/planes/w2_motor_pivotes.md.
"""
from __future__ import annotations

import json
from io import BytesIO

import pytest

pytest.importorskip("pydantic")
pytest.importorskip("sqlalchemy")
pytest.importorskip("fastapi")

import pandas as pd  # noqa: E402

from backend.schemas_pivot import PivotSpec  # noqa: E402
from backend.rgenerator.core.pivot_engine import pivot  # noqa: E402
from tests.factories import (  # noqa: E402
    auth_header_for,
    make_dimension,
    make_metric,
    make_metric_data,
    make_org,
    make_user,
)


# ─────────────────────────────────────────────────────────────────────────
# Datos de referencia (2 cursos × 2 meses × logro conocido)
# ─────────────────────────────────────────────────────────────────────────

# (Curso, Mes, Logro) — 6 filas; I A/Marzo tiene 2 obs para probar promedio.
_ROWS = [
    ("I A", "Marzo", 0.80),
    ("I A", "Marzo", 0.60),
    ("I A", "Abril", 0.90),
    ("I B", "Marzo", 0.50),
    ("I B", "Abril", 0.70),
    ("I B", "Abril", 0.30),
]

_PIVOT_SPEC = {
    "rows": ["Curso"],
    "cols": ["Mes"],
    "values": [{"field": "Logro", "agg": "mean", "format": ".1%"}],
    "totals": {"rows": True, "cols": True},
    "order": {"Mes": ["Marzo", "Abril"]},
}


def _reference_df() -> pd.DataFrame:
    return pd.DataFrame(_ROWS, columns=["Curso", "Mes", "Logro"])


@pytest.fixture
def metric_pivote(db_session, org):
    """Metric 'Logro' (float) + dims Curso/Mes con los datos de referencia."""
    dim_curso = make_dimension(db_session, org, name="Curso")
    dim_mes = make_dimension(db_session, org, name="Mes")
    metric = make_metric(
        db_session, org, name="Logro", data_type="float",
        dimensions=[dim_curso, dim_mes],
    )
    for curso, mes, val in _ROWS:
        make_metric_data(
            db_session, metric, value=str(val),
            dimensions_json={
                str(dim_curso.id_dimension): curso,
                str(dim_mes.id_dimension): mes,
            },
        )
    return metric


def _make_pivot_table_spec(db_session, org, *, metric_id, pivot_spec=None):
    """Crea un Spec tipo Tablas en modo pivote."""
    from backend.models import Spec
    cfg = {
        "version": 1,
        "data_source": {"metric_id": metric_id, "filters": {}, "derived_fields_override": []},
        "columns": [],
        "behavior": {},
        "pivot": pivot_spec or _PIVOT_SPEC,
    }
    spec = Spec(
        name="Tabla Pivote",
        type="Tablas",
        metadata_=json.dumps({"description": "", "is_draft": True}),
        charts_list="[]",
        tables_list=json.dumps([cfg]),
        org_id=org.id,
    )
    db_session.add(spec)
    db_session.commit()
    db_session.refresh(spec)
    return spec


@pytest.fixture
def tabla_pivote_creada(db_session, org, metric_pivote):
    return _make_pivot_table_spec(db_session, org, metric_id=metric_pivote.id_metric)


# ─────────────────────────────────────────────────────────────────────────
# Consumidor 1 — Dashboard: /data devuelve PivotResult
# ─────────────────────────────────────────────────────────────────────────


@pytest.mark.integration
class TestDashboardPivot:
    def test_sin_auth_401(self, client, tabla_pivote_creada):
        assert client.get(f"/api/tables/{tabla_pivote_creada.id_spec}/data").status_code == 401

    def test_data_devuelve_pivot_result(self, client_auth, tabla_pivote_creada):
        r = client_auth.get(f"/api/tables/{tabla_pivote_creada.id_spec}/data")
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["mode"] == "pivot"
        assert body["n_rows"] == 6
        pv = body["pivot"]
        # Estructura del PivotResult
        assert pv["row_fields"] == ["Curso"]
        assert pv["col_fields"] == ["Mes"]
        # 2 columnas de cuerpo (Marzo, Abril) + 1 Total
        labels = [c["keys"] for c in pv["columns"]]
        assert ["Marzo"] in labels and ["Abril"] in labels
        assert any(c["is_total"] for c in pv["columns"])

    def test_data_valores_correctos(self, client_auth, tabla_pivote_creada):
        r = client_auth.get(f"/api/tables/{tabla_pivote_creada.id_spec}/data")
        pv = r.json()["pivot"]
        # Fila I A: Marzo=mean(0.8,0.6)=0.7, Abril=0.9, Total=mean(0.8,0.6,0.9)
        rows = {tuple(row["keys"]): row for row in pv["rows"]}
        ia = rows[("I A",)]
        cols = [c["keys"][0] for c in pv["columns"]]
        # localizar índice de cada columna
        idx_marzo = cols.index("Marzo")
        idx_abril = cols.index("Abril")
        assert ia["cells"][idx_marzo]["value"] == pytest.approx(0.7)
        assert ia["cells"][idx_abril]["value"] == pytest.approx(0.9)
        assert ia["cells"][idx_marzo]["display"] == "70.0%"

    def test_multi_tenant_404(self, client, db_session, tabla_pivote_creada):
        # Usuario de otra org no puede leer la tabla
        other = make_org(db_session, name="Otra Org Pivot")
        other_user = make_user(db_session, other)
        r = client.get(
            f"/api/tables/{tabla_pivote_creada.id_spec}/data",
            headers=auth_header_for(other_user),
        )
        assert r.status_code == 404

    def test_preview_pivot(self, client_auth, metric_pivote):
        r = client_auth.post("/api/tables/preview", json={
            "config": {
                "data_source": {"metric_id": metric_pivote.id_metric, "filters": {}},
                "pivot": _PIVOT_SPEC,
            },
        })
        assert r.status_code == 200, r.text
        assert r.json()["mode"] == "pivot"


# ─────────────────────────────────────────────────────────────────────────
# Consumidor 2 — Excel: /export-pivot devuelve .xlsx válido
# ─────────────────────────────────────────────────────────────────────────


@pytest.mark.integration
class TestExcelPivot:
    def test_sin_auth_401(self, client, tabla_pivote_creada):
        assert client.get(f"/api/tables/{tabla_pivote_creada.id_spec}/export-pivot").status_code == 401

    def test_export_xlsx_valido(self, client_auth, tabla_pivote_creada):
        r = client_auth.get(f"/api/tables/{tabla_pivote_creada.id_spec}/export-pivot")
        assert r.status_code == 200, r.text
        assert r.headers["content-type"] == \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert r.content[:2] == b"PK"  # zip magic → xlsx

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        assert header[0] == "Curso"
        # Valores crudos + number_format porcentual
        data = {row[0]: row for row in rows[1:]}
        assert "I A" in data and "I B" in data and "Total" in data
        # Marzo I A = 0.7 (crudo, no "70.0%")
        marzo_idx = [i for i, h in enumerate(header) if "Marzo" in str(h)][0]
        assert data["I A"][marzo_idx] == pytest.approx(0.7)
        # number_format porcentual en una celda de datos
        cell = ws.cell(row=2, column=marzo_idx + 1)
        assert "%" in cell.number_format

    def test_export_400_si_no_pivote(self, client_auth, db_session, org, metric_pivote):
        # Tabla tabular clásica (sin pivot) → 400
        spec = _make_pivot_table_spec(db_session, org, metric_id=metric_pivote.id_metric)
        from backend.models import Spec
        s = db_session.get(Spec, spec.id_spec)
        cfg = json.loads(s.tables_list)[0]
        cfg.pop("pivot", None)
        s.tables_list = json.dumps([cfg])
        db_session.commit()
        r = client_auth.get(f"/api/tables/{spec.id_spec}/export-pivot")
        assert r.status_code == 400

    def test_multi_tenant_404(self, client, db_session, tabla_pivote_creada):
        other = make_org(db_session, name="Otra Org Excel")
        other_user = make_user(db_session, other)
        r = client.get(
            f"/api/tables/{tabla_pivote_creada.id_spec}/export-pivot",
            headers=auth_header_for(other_user),
        )
        assert r.status_code == 404


# ─────────────────────────────────────────────────────────────────────────
# Consumidor 3 — PDF v2: sección pivot en el runtime
# ─────────────────────────────────────────────────────────────────────────


@pytest.mark.integration
class TestPdfV2Pivot:
    def test_seccion_pivot_html(self, tmp_path):
        from backend.rgenerator.reports.runtime import _ejecutar_seccion
        seccion = {
            "tipo": "pivot", "titulo": "Pivote Logro", "df_input": "estudiantes",
            "spec": _PIVOT_SPEC,
        }
        out = _ejecutar_seccion(seccion, {"estudiantes": _reference_df()}, tmp_path)
        assert out["tipo"] == "table"
        assert "{{" not in out["html"] and "}}" not in out["html"]
        # Valores formateados presentes
        assert "70.0%" in out["html"]  # I A Marzo
        assert "Total" in out["html"]

    def test_seccion_pivot_con_filtro(self, tmp_path):
        from backend.rgenerator.reports.runtime import _ejecutar_seccion
        seccion = {
            "tipo": "pivot", "titulo": "Pivote I A", "df_input": "estudiantes",
            "spec": {"rows": ["Curso"], "cols": ["Mes"],
                     "values": [{"field": "Logro", "agg": "mean", "format": ".1%"}]},
            "filtro": {"Curso": "I A"},
        }
        out = _ejecutar_seccion(seccion, {"estudiantes": _reference_df()}, tmp_path)
        assert out["tipo"] == "table"
        assert "I A" in out["html"]
        assert "I B" not in out["html"]  # filtrado

    def test_construir_pdf_bytes(self, tmp_path):
        """Genera bytes PDF reales de un esquema con secciones pivote."""
        pytest.importorskip("weasyprint", reason="weasyprint no instalado en este entorno")
        from backend.rgenerator.reports import runtime
        # Crear un report_type temporal dentro de REPORTS_DIR con esquema pivote.
        rt_dir = runtime.REPORTS_DIR / "_pivot_test_tmp"
        rt_dir.mkdir(exist_ok=True)
        esquema = {
            "title": "Test Pivote",
            "secciones_fijas": [
                {"tipo": "pivot", "titulo": "Pivote A", "df_input": "estudiantes",
                 "spec": _PIVOT_SPEC},
                {"tipo": "pivot", "titulo": "Pivote B", "df_input": "estudiantes",
                 "spec": {"rows": ["Mes"], "values": [{"field": "Logro", "agg": "count"}]}},
            ],
        }
        (rt_dir / "esquema.json").write_text(json.dumps(esquema), encoding="utf-8")
        try:
            pdf = runtime.construir_pdf("_pivot_test_tmp", {"estudiantes": _reference_df()})
            assert isinstance(pdf, bytes)
            assert pdf.startswith(b"%PDF")
            assert b"{{" not in pdf
        finally:
            (rt_dir / "esquema.json").unlink(missing_ok=True)
            rt_dir.rmdir()


# ─────────────────────────────────────────────────────────────────────────
# Paridad — misma PivotSpec, mismos números en los tres consumidores
# ─────────────────────────────────────────────────────────────────────────


@pytest.mark.integration
class TestParidadConsumidores:
    def test_mismos_numeros_tres_consumidores(self, client_auth, tabla_pivote_creada):
        df = _reference_df()
        spec = PivotSpec(**_PIVOT_SPEC)
        engine_result = pivot(df, spec)

        # Referencia: {(row_key, col_key): value_crudo}
        ref = {}
        for row in engine_result.rows:
            rk = tuple(row.keys)
            for col, cell in zip(engine_result.columns, row.cells):
                ref[(rk, tuple(col.keys))] = cell.value

        # 1) Dashboard
        dash = client_auth.get(f"/api/tables/{tabla_pivote_creada.id_spec}/data").json()["pivot"]
        for row in dash["rows"]:
            rk = tuple(row["keys"])
            for col, cell in zip(dash["columns"], row["cells"]):
                assert cell["value"] == pytest.approx(ref[(rk, tuple(col["keys"]))]) \
                    if ref[(rk, tuple(col["keys"]))] is not None else cell["value"] is None

        # 2) Excel — valores crudos
        from openpyxl import load_workbook
        xls = client_auth.get(f"/api/tables/{tabla_pivote_creada.id_spec}/export-pivot")
        wb = load_workbook(BytesIO(xls.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        # columnas de datos (offset 1 por 'Curso')
        col_keys_xlsx = []
        for h in header[1:]:
            # "Marzo · Logro" → "Marzo"; "Total · Logro" → "Total"
            col_keys_xlsx.append(str(h).split(" · ")[0])
        for row in rows[1:]:
            rk = (row[0],)
            for j, ck in enumerate(col_keys_xlsx):
                expected = ref[(rk, (ck,))]
                got = row[j + 1]
                if expected is None:
                    assert got is None
                else:
                    assert got == pytest.approx(expected)

        # 3) PDF v2 — display formateado coincide con el del motor
        from backend.rgenerator.reports.tables import tabla_pivote
        df_pdf = tabla_pivote(df.copy(), spec=_PIVOT_SPEC)
        # I A Marzo display = "70.0%"
        ia = df_pdf[df_pdf["Curso"] == "I A"].iloc[0]
        assert ia["Marzo"] == "70.0%"
