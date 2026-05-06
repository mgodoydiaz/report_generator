"""Tests del step RunExcelETL.

Cobertura:
- Lectura básica con header_row.
- Lectura con metadata_cells (caso DIA: Establecimiento en B5, Curso en
  B6, datos desde fila 13).
- Combinación de metadata_cells + select_columns + rename_columns +
  enrich.
- Helpers _parse_a1 y _read_metadata_cells.
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest

from rgenerator.core.etl_steps import (
    RunExcelETL,
    _parse_a1,
    _read_metadata_cells,
)


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────

def _make_dia_xlsx(path: Path, establecimiento: str, curso: str,
                   estudiantes: list[dict], header_row_idx: int = 12) -> Path:
    """Genera un XLSX con shape DIA: metadata en B5/B6, header en fila
    `header_row_idx + 1` (1-indexed → fila 13 si idx=12), datos abajo.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    # Metadata: B5 (fila 5, col 2) = establecimiento, B6 = curso
    ws.cell(row=5, column=2, value=establecimiento)
    ws.cell(row=6, column=2, value=curso)
    # Header en fila (header_row_idx + 1) — pandas usa 0-indexed para
    # el header param, así que fila 13 visual = header=12.
    df = pd.DataFrame(estudiantes)
    header_row_visual = header_row_idx + 1
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row_visual, column=col_idx, value=col_name)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=header_row_visual + row_offset, column=col_idx, value=val)
    wb.save(path)
    return path


def _make_ctx(inputs: dict, params: dict | None = None):
    return SimpleNamespace(
        inputs=dict(inputs),
        artifacts={},
        params=dict(params or {}),
        user_inputs={},
        last_artifact_key=None,
        last_step=None,
    )


# ─────────────────────────────────────────────────────────────────────────
# Helpers internos: _parse_a1
# ─────────────────────────────────────────────────────────────────────────

class TestParseA1:
    def test_a1(self):
        assert _parse_a1("A1") == (0, 0)

    def test_b5(self):
        assert _parse_a1("B5") == (4, 1)

    def test_z1(self):
        assert _parse_a1("Z1") == (0, 25)

    def test_aa1(self):
        assert _parse_a1("AA1") == (0, 26)

    def test_lowercase_y_espacios(self):
        assert _parse_a1(" b5 ") == (4, 1)

    def test_invalido_lanza(self):
        with pytest.raises(ValueError):
            _parse_a1("5B")
        with pytest.raises(ValueError):
            _parse_a1("BB")


# ─────────────────────────────────────────────────────────────────────────
# Helpers internos: _read_metadata_cells
# ─────────────────────────────────────────────────────────────────────────

class TestReadMetadataCells:
    def test_lee_b5_y_b6(self, tmp_path):
        xlsx = _make_dia_xlsx(
            tmp_path / "curso_1A.xlsx",
            establecimiento="Colegio PHP",
            curso="1A",
            estudiantes=[
                {"Numero Lista": 1, "Nombre del Estudiante": "Juan",
                 "Pregunta 1": 80, "Pregunta 2": 90},
            ],
        )
        out = _read_metadata_cells(str(xlsx), [
            {"column_name": "Establecimiento", "cell": "B5"},
            {"column_name": "Curso", "cell": "B6"},
        ])
        assert out == {"Establecimiento": "Colegio PHP", "Curso": "1A"}

    def test_lista_vacia_retorna_dict_vacio(self):
        assert _read_metadata_cells("ignored.xlsx", []) == {}


# ─────────────────────────────────────────────────────────────────────────
# RunExcelETL con metadata_cells
# ─────────────────────────────────────────────────────────────────────────

class TestRunExcelETLMetadataCells:
    def test_dia_un_archivo(self, tmp_path):
        xlsx = _make_dia_xlsx(
            tmp_path / "curso_2B.xlsx",
            establecimiento="Liceo Pullinque",
            curso="2B",
            estudiantes=[
                {"Numero Lista": 1, "Nombre del Estudiante": "Ana",
                 "Pregunta 1": 100, "Pregunta 2": 80},
                {"Numero Lista": 2, "Nombre del Estudiante": "Luis",
                 "Pregunta 1": 60, "Pregunta 2": 70},
            ],
        )
        ctx = _make_ctx(
            inputs={"estudiantes": [str(xlsx)]},
            params={
                "header_row": 12,
                "metadata_cells": [
                    {"column_name": "Establecimiento", "cell": "B5"},
                    {"column_name": "Curso", "cell": "B6"},
                ],
            },
        )
        step = RunExcelETL(input_key="estudiantes")
        step.run(ctx)
        df = ctx.artifacts["df_consolidado_estudiantes"]
        assert len(df) == 2
        assert "Establecimiento" in df.columns
        assert "Curso" in df.columns
        assert (df["Establecimiento"] == "Liceo Pullinque").all()
        assert (df["Curso"] == "2B").all()
        assert "Pregunta 1" in df.columns

    def test_dia_consolida_varios_archivos_con_metadata_distinta(self, tmp_path):
        xlsx_1A = _make_dia_xlsx(
            tmp_path / "curso_1A.xlsx",
            establecimiento="Colegio PHP", curso="1A",
            estudiantes=[
                {"Numero Lista": 1, "Nombre del Estudiante": "Juan", "P1": 80},
            ],
        )
        xlsx_1B = _make_dia_xlsx(
            tmp_path / "curso_1B.xlsx",
            establecimiento="Colegio PHP", curso="1B",
            estudiantes=[
                {"Numero Lista": 1, "Nombre del Estudiante": "Pedro", "P1": 60},
                {"Numero Lista": 2, "Nombre del Estudiante": "Sofía", "P1": 90},
            ],
        )
        ctx = _make_ctx(
            inputs={"estudiantes": [str(xlsx_1A), str(xlsx_1B)]},
            params={
                "header_row": 12,
                "metadata_cells": [
                    {"column_name": "Establecimiento", "cell": "B5"},
                    {"column_name": "Curso", "cell": "B6"},
                ],
            },
        )
        RunExcelETL(input_key="estudiantes").run(ctx)
        df = ctx.artifacts["df_consolidado_estudiantes"]
        assert len(df) == 3
        # Cada fila tiene el curso correcto
        assert df[df["Nombre del Estudiante"] == "Juan"]["Curso"].iloc[0] == "1A"
        assert df[df["Nombre del Estudiante"] == "Pedro"]["Curso"].iloc[0] == "1B"
        assert df[df["Nombre del Estudiante"] == "Sofía"]["Curso"].iloc[0] == "1B"

    def test_metadata_cells_combinado_con_select_y_rename(self, tmp_path):
        xlsx = _make_dia_xlsx(
            tmp_path / "curso_3C.xlsx",
            establecimiento="X", curso="3C",
            estudiantes=[{"N°": 1, "Nombre del Estudiante": "Ana", "P1": 50, "P2": 50}],
        )
        ctx = _make_ctx(
            inputs={"estudiantes": [str(xlsx)]},
            params={
                "header_row": 12,
                "select_columns": ["Nombre del Estudiante", "P1"],
                "rename_columns": {"Nombre del Estudiante": "Nombre"},
                "metadata_cells": [
                    {"column_name": "Curso", "cell": "B6"},
                ],
            },
        )
        RunExcelETL(input_key="estudiantes").run(ctx)
        df = ctx.artifacts["df_consolidado_estudiantes"]
        # Select aplicó: solo P1 (no P2)
        assert "P2" not in df.columns
        # Rename aplicó: Nombre del Estudiante → Nombre
        assert "Nombre" in df.columns
        assert "Nombre del Estudiante" not in df.columns
        # metadata_cells añadió Curso
        assert df["Curso"].iloc[0] == "3C"

    def test_sin_metadata_cells_no_rompe_flujo_existente(self, tmp_path):
        xlsx = _make_dia_xlsx(
            tmp_path / "simple.xlsx",
            establecimiento="X", curso="1A",
            estudiantes=[{"a": 1, "b": 2}],
            header_row_idx=0,
        )
        ctx = _make_ctx(
            inputs={"data": [str(xlsx)]},
            params={"header_row": 0},
        )
        RunExcelETL(input_key="data").run(ctx)
        df = ctx.artifacts["df_consolidado_data"]
        # __source_file__ se inyecta automáticamente por RunExcelETL para
        # trazabilidad (ver pdf_steps_stub flow). El flujo legacy debe
        # seguir trabajando con las columnas del Excel original.
        assert [c for c in df.columns if c != "__source_file__"] == ["a", "b"]
        assert df.iloc[0]["a"] == 1
